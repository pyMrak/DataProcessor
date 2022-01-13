from openpyxl.styles import Font, Fill, PatternFill, NamedStyle, Border, Side
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.drawing.image import Image

import os

if os.path.dirname(__file__) == os.getcwd():
    import Paths
    import Graph
    import Basic
else:
    from . import Paths
    from . import Graph
    from . import  Basic

class ExcelWorkbook(Workbook):

    def __init__(self):
        Workbook.__init__(self)
        self._sheets = [ExcelWorksheet(self, "Sheet1")]

    def create_sheet(self, title=None, index=None):
        if title is None:
            title = "Sheet{}".format(len(self.worksheets)+1)
        self._sheets.append(ExcelWorksheet(self, title))

    def setSheetTitle(self, title):
        self.active.title = title

    def addBlock(self, block):
        self.active.addBlock(block)
        return block.dataGroup + 1

    # def reset(self):
    #     for sheet in self._sheets:
    #         sheet.reset()

    def save(self, filename, *args, **kwargs):
        for worksheet in self.worksheets:
            worksheet.insert(*args, **kwargs)
        self.active = 0
        super().save(filename)

class ExcelWorksheet(Worksheet):

    def __init__(self, parent, title):
        Worksheet.__init__(self, parent, title)
        self.blocks = []

    def addBlock(self, block):
        self.blocks.append(block)

    def insert(self, *args, **kwargs):
        kwargs["worksheet"] = self
        kwargs["occupied"] = [0, 0]
        for block in self.blocks:
            kwargs["occupied"] = block.insert(*args, **kwargs)

    # def reset(self):
    #     print("title:", self.title)
    #     Worksheet.__init__(self.parent, self.title)





class ExcelBlock(object):

    badVal = NamedStyle(name="bad")
    goodVal = NamedStyle(name="good")
    badVal.fill = PatternFill("solid", fgColor="ff8585")
    goodVal.fill = PatternFill("solid", fgColor="7beded")
    bd = Side(style='thin', color="000000")
    badVal.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    goodVal.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    numberFormats = ['0', '0.0', '0.00', '0.000', '0.0000']

    def __init__(self, position=None, direction=0, dataGroup=0):
        self.direction = direction
        if position is None:
            self.relPosition = [0, 0]
            self.position = [0, 0]
        else:
            self.definePosition(position)
        self.dataGroup = dataGroup
        #self.worksheet = None

    def definePosition(self, position):
        self.relPosition = position
        self.position = position


    def getPosition(self, *args, **kwargs):
        if "worksheet" in kwargs:
            self.worksheet = kwargs["worksheet"]
        else:
            return False
        if "occupied" in kwargs:
            if self.direction:
                self.position[0] = kwargs["occupied"][0] + self.relPosition[0]
                self.position[1] = self.relPosition[1]
            else:
                self.position[0] = self.relPosition[0]
                self.position[1] = kwargs["occupied"][1] + self.relPosition[1]
        return True

    def getOccupied(self, endPosition, **kwargs):
        if "occupied" in kwargs:
            if self.direction:
                return [kwargs["occupied"][0] + endPosition[0],
                        max(kwargs["occupied"][1], endPosition[1])]
            else:
                return [endPosition[0],
                        kwargs["occupied"][1] + endPosition[1]]

    def getAnchor(self, column, row):
        leftOver = column+1
        chars = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
        charsLen = len(chars)
        anchor = str(row+1)
        while leftOver > 0:
            mod = leftOver % charsLen
            char = chars[mod-1]
            leftOver = int((leftOver - mod)/charsLen)
            anchor = char + anchor
        return anchor



class ExcelPictureList(ExcelBlock):

    def __init__(self):
        self.path = None
        self.dataGroup = 0

    def setPath(self, path):
        self.path = path

    def setDataGroup(self, dg):
        self.dataGroup = dg


class ExcelParameterTable(ExcelBlock):

    def __init__(self, settings=None, position=None, direction=0, dataGroup=0):
        #self.dataGroup = dataGroup
        ExcelBlock.__init__(self, position, direction, dataGroup)

    def setDataGroup(self, dg):
        self.dataGroup = dg

    def insert(self, *args, **kwargs):
        if self.getPosition(*args, **kwargs) and "groups" in kwargs:
            dataGroup = kwargs["groups"][self.dataGroup]
            parameters = dataGroup["parameters"]
            self.worksheet.merge_cells(start_row=self.position[1]+1,
                                       start_column=self.position[0]+1,
                                       end_row=self.position[1]+1,
                                       end_column=self.position[0]+len(parameters)+1)
            cell = self.worksheet.cell(row=self.position[1] + 1,
                                       column=self.position[0] + 1)
            cell.value = parameters.groupName
            for row, index in enumerate(parameters.df.index):
                cell = self.worksheet.cell(row=self.position[1]+row+3,
                                           column=self.position[0]+1)
                cell.value = str(index)
                if True:# evaluate: TODO
                    if parameters.areValid(index):
                        cell.style = self.goodVal
                    else:
                        cell.style = self.badVal
            for i, entity in enumerate(parameters):
                col = self.position[0] + i + 2
                cell = self.worksheet.cell(row=self.position[1]+2,
                                           column=col)
                cell.value = entity
                rounding = parameters.getRounding(entity)
                goodStyle = self.goodVal.__copy__()
                goodStyle.name += "_" + entity
                goodStyle.number_format = self.numberFormats[rounding]
                badStyle = self.badVal.__copy__()
                badStyle.name += "_" + entity
                badStyle.number_format = self.numberFormats[rounding]
                for j, dPoint in enumerate(parameters[entity]):
                    cell = self.worksheet.cell(row=self.position[1]+j+3,
                                               column=col)
                    cell.value = dPoint.value
                    if True:  # evaluate: TODO
                        if parameters[j][i].isValid():
                            style = goodStyle#self.goodVal.__copy__()
                        else:
                            style = badStyle#self.badVal.__copy__()
                    #style.number_format = self.numberFormats[dPoint.rounding]
                    cell.style = style
                    #cell.style.number_format = self.numberFormats[dPoint.rounding]
            return self.getOccupied([col, self.position[1]+j+3], **kwargs)
        return self.getOccupied([0, 0], **kwargs)


class ExcelGraphList(ExcelBlock):
    graphSize = (19.20, 7.6)

    def __init__(self, settings=None, position=None, direction=0, dataGroup=0, listDown=True):
        self.setGraphFile(settings)
        #self.dataGroup = dataGroup
        self.listDown = listDown
        ExcelBlock.__init__(self, position, direction, dataGroup)

    def setGraphFile(self, graphFile):
        self._gw = Graph.GraphWrapper(graphFile)

    def setDataGroup(self, dg):
        self.dataGroup = dg

    def getGraphs(self, *args, **kwargs):
        graphs = []
        if "groups" in kwargs:
            dg = kwargs["groups"][self.dataGroup]
            cont = True
            graphs.append(self._gw.graph.draw(dg, draw=False, size=self.graphSize))
            while cont:
                cont = self._gw.graph.next()
                graphs.append(self._gw.graph.draw(dg, draw=False, size=self.graphSize))
        return graphs


    def insert(self, *args, **kwargs):
        if self.getPosition(*args, **kwargs) and "groups" in kwargs:
            graphs = self.getGraphs(*args, **kwargs)
            for i, graph in enumerate(graphs):
                img = Image(graph)
                img.anchor = self.getAnchor(self.position[0]+i*30*(not self.listDown),
                                            self.position[1]+i*38*self.listDown)
                self.worksheet.add_image(img)
            return self.getOccupied([self.position[0]+i*30*(not self.listDown)+30,
                                     self.position[1]+i*38*self.listDown+38], **kwargs)
        return self.getOccupied([0, 0], **kwargs)


class Report(object):

    def __init__(self):
        self.lastReport = None

    def openLast(self):
        if self.lastReport is not None:
            if type(self.lastReport) == list:
                for path in self.lastReport:
                    self.openReport(path)
            elif type(self.lastReport) == str:
                self.openReport(self.lastReport)

    def openReport(self, path):
        os.startfile(path.replace('/', '\\'), 'open')




class ExcelReport(Report):

    BLOCKS = {
        "ExcelGraphList": ExcelGraphList,
        "ExcelParameterTable": ExcelParameterTable,
    }

    def __init__(self):
        Report.__init__(self)
        self.workbook = ExcelWorkbook()
        self.nameRaw = "{}"
        self.maxDataGroups = 0

    def readReportSetting(self, settingsFile, GUIobj=None):
        self.maxDataGroups = 0
        self.settingsFile = settingsFile
        self.settings = Basic.loadUserExcReportFile(settingsFile, GUIobj)
        if "name" in self.settings:
            self.nameRaw = self.settings["name"]
        if "sheets" in self.settings:
            sheets = self.settings["sheets"]
            for i, sheetName in enumerate(sheets):
                if i:
                    self.workbook.create_sheet(sheetName)
                    self.workbook.active = i
                else:
                    self.workbook.setSheetTitle(sheetName)
                if "blocks" in sheets[sheetName]:
                    blocks = sheets[sheetName]["blocks"]
                    for block in blocks:
                        if block in self.BLOCKS:
                            dataGroup = self.workbook.addBlock(self.BLOCKS[block](**blocks[block]))
                            if dataGroup > self.maxDataGroups:
                                self.maxDataGroups = dataGroup
        return self.maxDataGroups

    def reset(self):
        self.workbook = ExcelWorkbook()
        self.nameRaw = "{}"
        self.maxDataGroups = 0
        self.readReportSetting(self.settingsFile)

    def save(self, dataGroups, path=None):
        if len(dataGroups) < self.maxDataGroups:
            return None  # TODO add error
        fileName = self.nameRaw.format(dataGroups[0].getFolder())
        if path is None:
            print(dataGroups[0].getPath(), fileName)
            path = Paths.join(dataGroups[0].getPath(), fileName)
        else:
            path = Paths.join(path, fileName)
        print("saving")
        for i in range(100):
            if i:
                nr = "({})".format(i)
            else:
                nr = ""
            try:
                filePath = "{0}{1}.xlsx".format(path, nr)
                self.workbook.save(filePath, groups=dataGroups)
                self.reset()
                self.lastReport = filePath
                return filePath
            except PermissionError:
                self.reset()







if __name__ == "__main__":
    from Data import DataGroup

    # wb = ExcelWorkbook()
    # wb.setSheetTitle("Parametri")
    # #wb.adBlock(ExcelParameterTable([2, 3]))
    #
    # # wb.create_sheet("Grafi")
    # # wb.active = 1
    # wb.adBlock(ExcelGraphList('functionalGP', [2, 3], 0))
    # wb.adBlock(ExcelParameterTable([1, 1], 0))
    # # wb.adBlock(ExcelParameterTable([2, 3], 0))
    er = ExcelReport()
    er.readReportSetting("functional report")



    data = DataGroup()
    data.setFolder(Paths.testDir)
    data.readPyro('functionalLab')
    # sfg = SeriesFunctionGroup()
    # sfg.load('5011-721-414')
    # sfg.apply(data)
    data.getParameters('5011-721-414-QC')

    #wb.save("test.xlsx", groups=[data])
    er.save([data])
    er.openLast()
    # eg = ExcelGraph('functionalGP', [2, 3], 0)
    # print(eg.getAnchor(27, 5))


